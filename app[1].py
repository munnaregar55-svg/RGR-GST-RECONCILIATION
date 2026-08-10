from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
import os, io, uuid

app=Flask(__name__)
app.secret_key="rgr-secret"
UPLOAD_DIR="data/uploads"; REPORT_DIR="data/reports"
os.makedirs(UPLOAD_DIR,exist_ok=True); os.makedirs(REPORT_DIR,exist_ok=True)

REQUIRED=["GSTIN","BILL NO","BILL AMOUNT","TAXABLE AMOUNT","IGST","CGST","SGST"]

def money(v):
    try: return round(float(v or 0),2)
    except: return 0.0

def read_sheet(path):
    wb=load_workbook(path,data_only=True,read_only=True)
    ws=wb.active; rows=list(ws.iter_rows(values_only=True))
    if not rows: return [], REQUIRED[:]
    headers=[str(x or "").strip().upper() for x in rows[0]]
    missing=[h for h in REQUIRED if h not in headers]
    data=[]
    for row in rows[1:]:
        if not any(x not in (None,"") for x in row): continue
        d={h:(row[i] if i<len(row) else None) for i,h in enumerate(headers) if h}
        data.append(d)
    return data,missing

def normalize(row):
    return {
      "gstin":str(row.get("GSTIN") or "").strip().upper().replace(" ",""),
      "invoice":str(row.get("BILL NO") or "").strip().upper().replace(" ",""),
      "party":str(row.get("PARTY NAME") or row.get("PARTY") or row.get("NAME") or "").strip(),
      "bill_amount":money(row.get("BILL AMOUNT")),
      "taxable":money(row.get("TAXABLE AMOUNT")),
      "igst":money(row.get("IGST")),
      "cgst":money(row.get("CGST")),
      "sgst":money(row.get("SGST"))
    }

def key(r): return (r["gstin"],r["invoice"])

def reason(d):
    names={"bill_amount":"Bill Amount","taxable":"Taxable Amount","igst":"IGST","cgst":"CGST","sgst":"SGST"}
    bad=[names[k] for k,v in d.items() if abs(v)>1]
    return "इनमें अंतर है: "+", ".join(bad)

def reconcile(books,gstr):
    b=[normalize(x) for x in books]; g=[normalize(x) for x in gstr]
    bm={key(x):x for x in b if key(x)!=("","")}
    gm={key(x):x for x in g if key(x)!=("","")}
    result={"matched":[],"mismatch":[],"not_2b":[],"not_books":[]}
    fields=["bill_amount","taxable","igst","cgst","sgst"]
    for k,br in bm.items():
        if k not in gm: result["not_2b"].append(br); continue
        gr=gm[k]; diffs={f:round(br[f]-gr[f],2) for f in fields}
        if any(abs(v)>1 for v in diffs.values()):
            result["mismatch"].append({"book":br,"gstr":gr,"diff":diffs,"reason":reason(diffs)})
        else: result["matched"].append({"book":br,"gstr":gr})
    for k,gr in gm.items():
        if k not in bm: result["not_books"].append(gr)
    return result

def export_rows(r):
    out=[]
    for x in r["mismatch"]:
        b=x["book"]; g=x["gstr"]; d=x["diff"]
        out.append(["राशि मैच नहीं हुई",b["party"] or g["party"],b["gstin"],b["invoice"],x["reason"],
          b["bill_amount"],g["bill_amount"],d["bill_amount"],b["taxable"],g["taxable"],d["taxable"],
          b["igst"],g["igst"],d["igst"],b["cgst"],g["cgst"],d["cgst"],b["sgst"],g["sgst"],d["sgst"]])
    for x in r["not_2b"]:
        out.append(["GSTR-2B में बिल नहीं मिला",x["party"],x["gstin"],x["invoice"],
          "Books में है, GSTR-2B में नहीं मिला",x["bill_amount"],"",x["bill_amount"],
          x["taxable"],"",x["taxable"],x["igst"],"",x["igst"],x["cgst"],"",x["cgst"],x["sgst"],"",x["sgst"]])
    for x in r["not_books"]:
        out.append(["Books में बिल नहीं मिला",x["party"],x["gstin"],x["invoice"],
          "GSTR-2B में है, Books में नहीं मिला","",x["bill_amount"],-x["bill_amount"],
          "",x["taxable"],-x["taxable"],"",x["igst"],-x["igst"],"",x["cgst"],-x["cgst"],"",x["sgst"],-x["sgst"]])
    return out

@app.route("/")
def index(): return render_template("index.html")

@app.route("/reconcile",methods=["POST"])
def do_reconcile():
    bf=request.files.get("books"); gf=request.files.get("gstr2b")
    if not bf or not gf or not bf.filename or not gf.filename:
        flash("❌ Books और GSTR-2B दोनों Excel files upload करें."); return redirect(url_for("index"))
    bp=os.path.join(UPLOAD_DIR,secure_filename(bf.filename)); gp=os.path.join(UPLOAD_DIR,secure_filename(gf.filename))
    bf.save(bp); gf.save(gp)
    try:
        books,bmiss=read_sheet(bp); gstr,gmiss=read_sheet(gp)
        if bmiss or gmiss:
            msg=[]
            if bmiss: msg.append("Books में missing: "+", ".join(bmiss))
            if gmiss: msg.append("GSTR-2B में missing: "+", ".join(gmiss))
            flash("❌ Required heading missing — "+" | ".join(msg)); return redirect(url_for("index"))
        result=reconcile(books,gstr)
    except Exception as e:
        flash("❌ Excel पढ़ने में समस्या: "+str(e)); return redirect(url_for("index"))
    rid=uuid.uuid4().hex[:10]
    import json
    with open(os.path.join(REPORT_DIR,rid+".json"),"w",encoding="utf8") as f: json.dump(result,f)
    return render_template("dashboard.html",result=result,rid=rid)

@app.route("/export/<rid>")
def export(rid):
    import json
    path=os.path.join(REPORT_DIR,rid+".json")
    if not os.path.exists(path): return "Report not found",404
    with open(path,encoding="utf8") as f: r=json.load(f)
    wb=Workbook(); ws=wb.active; ws.title="RGR Reconciliation"
    ws.append(["स्थिति","पार्टी नाम","GSTIN","Bill No","क्यों मैच नहीं हुआ","Books Bill Amount","2B Bill Amount","Bill Amount अंतर",
      "Books Taxable","2B Taxable","Taxable अंतर","Books IGST","2B IGST","IGST अंतर","Books CGST","2B CGST","CGST अंतर","Books SGST","2B SGST","SGST अंतर"])
    for row in export_rows(r): ws.append(row)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name="RGR_GST_Reconciliation.xlsx",
      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__=="__main__": app.run(debug=True,host="0.0.0.0",port=5000)
